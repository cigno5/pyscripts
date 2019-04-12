import argparse
import csv
import os

import xlrd
from openpyxl import load_workbook, Workbook


def excel_to_csv(excel_file, csv_file):
    with xlrd.open_workbook(excel_file) as wb, \
            open(csv_file, 'w') as your_csv_file:

        wr = csv.writer(your_csv_file,
                        quoting=csv.QUOTE_ALL,
                        delimiter=args.delimiter,
                        quotechar=args.quote_char)

        sh = wb.sheet_by_index(0)

        for rownum in range(sh.nrows):
            wr.writerow(sh.row_values(rownum))


def xls_split():
    xls_source = load_workbook(args.origin, read_only=True)
    sheet = xls_source[xls_source.sheetnames[0]]

    col_value_map = {}

    header_rows = []
    perc_step = int(sheet.max_row / 15)

    print('Splitting...')
    c = cp = 0
    for row in sheet.rows:
        c += 1
        cp += 1
        _progress = (c * 100) / sheet.max_row
        if cp == perc_step:
            print("  %.1f%% done..." % _progress, end='\r', flush=True)
            cp = 0

        current_row = [c.value for c in row]
        if c <= args.header:
            header_rows.append(current_row)
        else:
            cell_value = current_row[args.column]
            if cell_value not in col_value_map:
                _wb = Workbook(write_only=True)
                _ws = _wb.create_sheet()
                for _header_row in header_rows:
                    _ws.append(_header_row)

                col_value_map[cell_value] = (_wb, _ws)

            ws = col_value_map[cell_value][1]
            ws.append(current_row)

    print("Split done\nSaving...")
    for value, wb in [(it[0], it[1][0]) for it in col_value_map.items()]:
        new_file = "%s-%s%s" % (basename_ext[0], value, basename_ext[1])
        print("  %s..." % new_file)
        wb.save(os.path.join(output_dir, new_file))

    print("Done")


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("--delimiter", default=';', help="CSV column delimiter")
    parser.add_argument("--quote-char", default='"', help="CSV Quoting character")

    subparsers = parser.add_subparsers()

    xls2csv_parser = subparsers.add_parser("xls2csv")
    xls2csv_parser.set_defaults(command='xls2csv')
    xls2csv_parser.add_argument("excel", help="Excel file")
    xls2csv_parser.add_argument("csv", help="CSV destination")
    xls2csv_parser.add_argument("--overwrite", action="store_true", help="Delete output file if already exist")

    xlssplit_parser = subparsers.add_parser("split")
    xlssplit_parser.set_defaults(command='xls-split')
    xlssplit_parser.add_argument("origin", help="Original excel file")
    xlssplit_parser.add_argument("-c", "--column", type=int, help="Column to inspect", required=True)
    xlssplit_parser.add_argument("-p", "--prefix", help="Resulting file prefix")
    xlssplit_parser.add_argument("--header", type=int, help="Header row", default=1)
    xlssplit_parser.add_argument("-o", "--output-dir", help="Output directory (default same as origin)")

    args = parser.parse_args()

    if not hasattr(args, 'command'):
        parser.print_usage()
        exit(1)

    if args.command == 'xls2csv':
        assert args.excel, "Missing excel file"
        assert os.path.exists(args.excel), "Cannot find excel file"
        assert args.csv, "Missing csv file"
        if os.path.exists(args.csv):
            if args.overwrite:
                os.remove(args.csv)
            else:
                raise ValueError("CSV file already exists")

        excel_to_csv(args.excel, args.csv)

    elif args.command == 'xls-split':
        assert os.path.exists(args.origin)

        output_dir = args.output_dir or os.path.dirname(args.origin)
        os.makedirs(output_dir, exist_ok=True)
        basename_ext = os.path.splitext(os.path.basename(args.origin))

        xls_split()

    print("Done")
