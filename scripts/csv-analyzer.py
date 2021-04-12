import argparse
import csv
import datetime
import os.path
import re
import sys

import tabulate
from openpyxl import load_workbook

filter_re = re.compile("(\\d+)(=|!=)(.+)")

PROGRESS_STEP = 500


class DataReader:
    def __init__(self):
        self.current_line = -1
        self.returned_rows = 0

    def __iter__(self):
        return self

    def __next__(self):
        if args.limit and self.returned_rows >= args.limit:
            raise StopIteration

        if highest_line and self.current_line >= highest_line:
            raise StopIteration

        while True:
            try:
                self.current_line += 1

                row = self._next_row()

                # the header is not affected by filters
                if self.current_line > 0:
                    if lowest_line and self.current_line < lowest_line:
                        continue

                    if lines and self.current_line not in lines:
                        continue

                    if filters and not all([f(row) for f in filters]):
                        continue

                self.returned_rows += 1
                return self.current_line, row
            finally:
                self.print_progress()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def print_progress(self):
        if args.progress:
            print("Processed %i rows..." % (self.current_line + 1), end="\r", flush=True)

    def _next_row(self):
        raise NotImplemented


class XlsDataReader(DataReader):
    def __init__(self, file, sheet_num):
        super().__init__()

        self.xls_source = load_workbook(file, read_only=True)
        sheet = self.xls_source[self.xls_source.sheetnames[sheet_num]]

        self.__rows = sheet.rows
        self.__n_rows = sheet.max_row
        self.__local_progress = 0

    def _next_row(self):
        self.__local_progress += 1
        row = next(self.__rows)
        return [c.value for c in row]

    def __exit__(self, exc_type, exc_val, exc_tb):
        super().__exit__(exc_type, exc_val, exc_tb)
        self.xls_source.close()

    def print_progress(self):
        _progress = ((self.current_line + 1) * 100) / self.__n_rows
        if self.__local_progress == PROGRESS_STEP:
            print("  %.1f%% done..." % _progress, end="\r", flush=True)
            self.__local_progress = 0


class CsvDataReader(DataReader):
    def __init__(self, file, delimiter, quote_char):
        super().__init__()

        self.file_path = file
        self.file = open(file, 'r')
        self.data = csv.reader(self.file, delimiter=delimiter, quotechar=quote_char)
        self.__local_progress = 0
        self.__file_size = os.stat(file).st_size
        self.__estimated_n_rows = None

    def _next_row(self):
        self.__local_progress += 1

        row = next(self.data)

        if self.__estimated_n_rows is None:
            with open(self.file_path, 'r') as x:
                _len = 0
                _steps = 1000
                for _ in range(0, _steps):
                    _len += len(x.readline())
                _len /= _steps
                self.__estimated_n_rows = self.__file_size / _len

        return row

    def print_progress(self):
        _progress = ((self.current_line + 1) * 100) / self.__estimated_n_rows
        if self.__local_progress > PROGRESS_STEP:
            # print(".", end="", flush=True)
            print("  roughly %.1f%% done..." % _progress, end="\r", flush=True)
            self.__local_progress = 0

    def __exit__(self, exc_type, exc_val, exc_tb):
        super().__exit__(exc_type, exc_val, exc_tb)
        self.file.close()


class ColMetadata:
    def __init__(self, name, idx):
        self.name = name
        self.idx = idx
        self.first_value = None
        self.max_size = None
        self.max_size_row_idx = None
        self.max_size_value = None
        self.none_values_count = 0
        self.values_count = 0
        self.values = set()
        self.__guess_map = {}

    def push_value(self, row_idx, value):
        if value is None or value == "":
            self.none_values_count += 1
        else:
            if self.first_value is None:
                self.first_value = value

            str_value = str(value).strip()
            str_len = len(str_value)
            if self.max_size is None or self.max_size < str_len:
                self.max_size = str_len
                self.max_size_row_idx = row_idx
                self.max_size_value = str_value

            gt = self.__guess_type(value)
            if gt not in self.__guess_map:
                self.__guess_map[gt] = 0
            self.__guess_map[gt] += 1

            self.values.add(value)
            self.values_count += 1

    def guessed_type(self):
        popular_types = [t for t, c in sorted(self.__guess_map.items(), key=lambda x: x[1], reverse=True) if c > 0]
        return popular_types[0] if popular_types else None

    def colnum_letter(self):
        string = ""
        n = self.idx + 1
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def distinct_values(self):
        return len(self.values)

    @staticmethod
    def __guess_type(v):
        def is_int():
            return type(v) == int or float(int(v)) == float(v)

        def is_float():
            return type(v) == float or float(v)

        def is_datetime():
            return type(v) == datetime or datetime.datetime.strptime(v, "%Y-%m-%d %H:%M:%S")

        def is_string():
            return type(v) == str or str(v)

        guessers = [
            ("Integer", is_int),
            ("BigDecimal", is_float),
            ("Date", is_datetime),
            ("String", is_string),
        ]

        for g_type, g_func in guessers:
            try:
                if g_func():
                    return g_type
            except:
                continue

        return None


def enumerate_cells(row, consider_columns=True):
    _cells = [(i, val) for i, val in enumerate(row) if consider_columns is False or columns is None or i in columns]
    if consider_columns and columns:
        _sort_map = {c: x for x, c in enumerate(columns)}
        _cells.sort(key=lambda cell: _sort_map[cell[0]])

    return _cells


def chop_text(text):
    if not args.extended_text and text and len(str(text).strip()) > 50:
        return text[0:50] + " [...]"
    else:
        return text


def get_columns_meta():
    _idx_map = {}
    _, headers = next(data)

    col_map = [ColMetadata(header, i) for i, header in enumerate_cells(headers)]

    def get_col_map(x):
        if x not in _idx_map:
            for j, c in enumerate(col_map):
                if c.idx == x:
                    _idx_map[x] = j
                    break

        return col_map[_idx_map[x]]

    # collecting data
    for line, row in data:
        if args.limit and line > args.limit:
            break

        for i, value in enumerate_cells(row):
            get_col_map(i).push_value(line, value)

    return col_map


def analyze_columns():
    # creating output table
    table = [["Index", "Col", "Header",
              "First sample value",
              "Type (guess)",

              "Max size",
              "Max size line",
              "Max size value",

              "# Distinct values"]]

    for meta in get_columns_meta():
        table.append([meta.idx, meta.colnum_letter(), meta.name,
                      chop_text(meta.first_value),
                      meta.guessed_type(),

                      meta.max_size,
                      meta.max_size_row_idx,
                      chop_text(meta.max_size_value),

                      meta.distinct_values()])

    print_table(table)


def show_distinct_values():
    _values_map = {}

    def _extract_values_row():
        def get_values(c):
            if c.idx not in _values_map:
                _values_map[c.idx] = sorted((str(v) for v in c.values))
            return _values_map[c.idx]

        return [get_values(c)[x] if x < len(get_values(c)) else None for c in columns_meta]

    columns_meta = [c for c in get_columns_meta()]

    max_values_count = max([len(c.values) for c in columns_meta])
    table = [
        [''] + [c.name for c in columns_meta],
        ['Col'] + [c.colnum_letter() for c in columns_meta],
        ['Index'] + [c.idx for c in columns_meta],
        [''] + ['---------' for _ in columns_meta]
    ]

    first = True
    for x in range(0, max_values_count):
        table.append(['Values' if first else None] + _extract_values_row())
        first = False

    print_table(table)


def check_uniqueness():
    key_counter = {}

    next(data)  # skip header

    for line, row in data:
        key = "-".join([str(row[c]) for c in columns])
        if key in key_counter:
            key_counter[key].append(line)
        else:
            key_counter[key] = [line]

    non_unique_keys = dict([(k, c) for (k, c) in key_counter.items() if len(c) > 1])
    if non_unique_keys:
        print("Some values are not unique (%d tuples): " % len(non_unique_keys))
        table = [["Key", "# occ", "Lines #"]]
        too_much_counter = 0
        for k, v in non_unique_keys.items():
            too_much_counter += 1
            table.append([k, len(v), ", ".join((str(_v) for _v in v))])

            if too_much_counter > 50:
                table.append(['*** TOO MANY RESULTS ***', '***', "***"])
                break

        print_table(table)
    else:
        print("Column tuple is unique")


def show_rows():
    if args.hide_linenum:
        headers = [h for _, h in enumerate_cells(next(data)[1])]
    else:
        headers = ["line #"] + [h for _, h in enumerate_cells(next(data)[1])]

    table = [headers]

    c = 0
    for line, row in data:
        c += 1
        if args.limit and c > args.limit:
            break

        if args.hide_linenum:
            table.append([v for _, v in enumerate_cells(row)])
        else:
            table.append([line] + [v for _, v in enumerate_cells(row)])

    print_table(table)


def print_table(table):
    if args.format == 'csv':
        writer = csv.writer(sys.stdout,
                            delimiter=args.delimiter,
                            quotechar=args.quote_char,
                            quoting=csv.QUOTE_NONNUMERIC,
                            lineterminator='\n')
        for r in table:
            writer.writerow(r)
    elif args.format == 'human':
        print(tabulate.tabulate(table))
    else:
        print(str(table))


def new_reader():
    _ext = os.path.splitext(args.csv_file)[1][1:].lower()
    if _ext in ['csv', 'txt']:
        _reader = CsvDataReader(args.csv_file, args.delimiter, args.quote_char)
    elif _ext in ['xls', 'xlsx', 'xlsm']:
        _reader = XlsDataReader(args.csv_file, args.sheet_num - 1)
    else:
        raise ValueError('file %s is not valid' % args.csv_file)

    # skip rows
    for _ in range(0, args.skip_rows):
        next(_reader)

    return _reader


def __build_int_list(_ints):
    if _ints:
        list_ = []
        for _int in _ints.split(","):
            if '-' in _int:
                _from, _to = _int.split("-")
                list_ += range(int(_from), int(_to) + 1)
            else:
                list_.append(int(_int))

        return list_
    else:
        return None


def __build_filters(flts_):
    if flts_:
        def generate_filter(f):
            match = filter_re.match(f)
            if not match:
                raise ValueError("Filter '%s' is not valid" % f)
            col = match.group(1)
            filter_negated = match.group(2) == '!='
            text = match.group(3).strip().lower()

            def _filter(row):
                v = row[int(col.strip())]
                v = str(v).strip().lower()
                if text == '*empty':
                    ret = v is None or v == ''
                else:
                    ret = text in v

                if filter_negated:
                    ret = not ret

                return ret

            return _filter

        return [generate_filter(f) for f in flts_]
    else:
        return None


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("csv_file", help="CSV file path")
    parser.add_argument("--analyze", action="store_true", help="Analyze columns")
    parser.add_argument("--show", action="store_true", help="Show rows value")
    parser.add_argument("--distinct", action="store_true", help="Show distinct values")
    parser.add_argument("--unique", action="store_true", help="Check uniqueness")

    processing_options_grp = parser.add_argument_group("Processing options")
    processing_options_grp.add_argument("-c", "--columns",
                                        help="Columns indexes to be processed. Index starts from 0, list columns "
                                             "separated by comma or with range x-y (inclusive)")
    processing_options_grp.add_argument("-l", "--lines",
                                        help="Line number to be processed; separate by comma or range x-y (inclusive)")

    processing_options_grp.add_argument("-f", "--filter", nargs="*",
                                        help="Filter rows in dataset for processing. Examples:  "
                                             "1) --filter 0=text --> column 0 contains 'text'; "
                                             "2) --filter 0!=text --> column 0 doesn't contain 'text'; "
                                             "3) --filter 0=*empty --> column 0 is empty; "
                                             "4) --filter 0!=*empty --> column 0 is NOT empty; "
                                             "")
    processing_options_grp.add_argument("--limit", type=int, help="Limit the number of rows to be taken in evaluation")
    processing_options_grp.add_argument("--progress", action='store_true', help="Show processing progress")
    processing_options_grp.add_argument("--skip-rows", default=0, type=int,
                                        help="Define how many rows to be skipped before considering the file as started")

    output_options_grp = parser.add_argument_group("Output options")
    output_options_grp.add_argument("--format", choices=["csv", "human"], default="human", help="Output format")
    output_options_grp.add_argument("--extended-text", action='store_true',
                                    help="Show extended texts (by default text is cut after 50 chars)")
    output_options_grp.add_argument("--hide-linenum", action='store_true',
                                    help="Hide line number when show rows")

    csv_options_grp = parser.add_argument_group("CSV options")
    csv_options_grp.add_argument("--delimiter", default=';', help="CSV column delimiter")
    csv_options_grp.add_argument("--quote-char", default='"', help="CSV Quoting character")

    xls_options_grp = parser.add_argument_group("XLS Options")
    xls_options_grp.add_argument("--sheet-num", default=1, type=int, help="The excel's sheet number")

    args = parser.parse_args()

    assert os.path.exists(args.csv_file), "File %s doesn't exist" % args.csv_file

    # build up the list of columns to be processed
    columns = __build_int_list(args.columns)

    # build up the list of lines to be processed
    lines = __build_int_list(args.lines)
    lowest_line = min(lines) if lines else None
    highest_line = max(lines) if lines else None

    # build up the filter functions
    filters = __build_filters(args.filter)

    if args.analyze:
        with new_reader() as data:
            analyze_columns()
            print()

    if args.show:
        with new_reader() as data:
            show_rows()
            print()

    if args.distinct:
        with new_reader() as data:
            show_distinct_values()
            print()

    if args.unique:
        with new_reader() as data:
            check_uniqueness()
            print()
